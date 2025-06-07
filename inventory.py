# inventory.py
import os
from openpyxl import load_workbook

INVENTORY_FILE = "clime_db.xlsx"

def has_record(query: str) -> bool:
    """
    Checks if a given artist/album name exists in the inventory.
    Match is case-insensitive and substring-based.
    """
    if not os.path.exists(INVENTORY_FILE):
        return False

    wb = load_workbook(INVENTORY_FILE)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    if "Artist - Album" not in headers:
        return False

    name_idx = headers.index("Artist - Album")

    query = query.strip().lower()

    for row in ws.iter_rows(min_row=2, values_only=True):
        record_name = str(row[name_idx]).lower()
        if query in record_name:
            return True

    return False
