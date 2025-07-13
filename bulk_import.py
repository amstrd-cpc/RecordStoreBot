"""Bulk import records from Excel using Discogs lookup."""

import os
import sys
import tempfile
from typing import List

from openpyxl import load_workbook
from telegram import Update
from telegram.ext import (
    CommandHandler,
    MessageHandler,
    ConversationHandler,
    ContextTypes,
    filters,
)

from auth import require_auth
from add_record import quick_add_from_discogs
from db import init_db


WAITING_FOR_FILE = 0


def bulk_import_from_excel(path: str) -> int:
    """Read an Excel file and import records using Discogs lookup."""
    wb = load_workbook(path)
    ws = wb.active
    rows: List[str] = [cell[0] for cell in ws.iter_rows(values_only=True)]

    added = 0
    for i in range(0, len(rows), 2):
        title = rows[i]
        if title is None:
            continue
        price = rows[i + 1] if i + 1 < len(rows) else None
        if price is None:
            continue
        try:
            price_val = float(price)
        except (ValueError, TypeError):
            continue

        try:
            quick_add_from_discogs(str(title), price_val)
            added += 1
        except Exception as e:
            print(f"Failed to add '{title}': {e}")

    return added


@require_auth
async def start_bulk_import(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Prompt user to send an Excel file for bulk import."""
    await update.message.reply_text(
        "📥 Send the Excel file (two rows per record: title then price)."
    )
    return WAITING_FOR_FILE


async def handle_excel_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Receive the Excel file and process the import."""
    document = update.message.document
    if not document:
        await update.message.reply_text("❌ Please send an Excel file.")
        return WAITING_FOR_FILE

    file = await document.get_file()
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        await file.download_to_drive(tmp.name)
        temp_path = tmp.name

    await update.message.reply_text("⏳ Importing records...")

    count = bulk_import_from_excel(temp_path)
    os.remove(temp_path)

    await update.message.reply_text(f"✅ Imported {count} record(s) from file.")
    return ConversationHandler.END


async def cancel_bulk_import(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Cancel the bulk import conversation."""
    await update.message.reply_text("✅ Bulk import cancelled.")
    return ConversationHandler.END


def start_bulk_import_flow() -> ConversationHandler:
    """Return the ConversationHandler for the /bulkimport command."""
    return ConversationHandler(
        entry_points=[CommandHandler("bulkimport", start_bulk_import)],
        states={
            WAITING_FOR_FILE: [MessageHandler(filters.Document.ALL, handle_excel_file)]
        },
        fallbacks=[CommandHandler("cancel", cancel_bulk_import)],
        name="bulk_import",
        persistent=False,
    )


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python bulk_import.py <excel_file>")
        sys.exit(1)

    init_db()
    total = bulk_import_from_excel(sys.argv[1])
    print(f"Imported {total} record(s)")


