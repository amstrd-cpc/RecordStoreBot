# sales.py
import os
import datetime
from telegram import Update, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.ext import (
    CommandHandler, CallbackQueryHandler, MessageHandler,
    ContextTypes, ConversationHandler, filters
)
from inventory import load_inventory

from openpyxl import Workbook, load_workbook

INVENTORY_FILE = "clime_db.xlsx"
SALES_FOLDER = "sales"

if not os.path.exists(SALES_FOLDER):
    os.makedirs(SALES_FOLDER)

SEARCH_VINYL, CONFIRM_CART, CHOOSE_PAYMENT = range(3)

# 🧠 Conversation storage
user_sessions = {}

# 🛒 Start /sell
async def start_sell(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["cart"] = []
    context.user_data["page"] = 0
    context.user_data["inventory"] = load_inventory()
    await show_inventory_page(update, context)
    return SEARCH_VINYL

# 📃 Show inventory with pagination
async def show_inventory_page(update: Update, context: ContextTypes.DEFAULT_TYPE):
    page = context.user_data["page"]
    inventory = context.user_data["inventory"]
    items = inventory[page*10:(page+1)*10]

    if not items:
        await update.message.reply_text("No items found.")
        return

    buttons = []
    for i, item in enumerate(items):
        btn_text = f"{item['Artist - Album']} - ${item['USD Price']}"
        buttons.append([
            InlineKeyboardButton(btn_text[:60], callback_data=f"add_{page*10 + i}")
        ])

    nav = []
    if page > 0:
        nav.append(InlineKeyboardButton("⬅️ Prev", callback_data="prev"))
    if len(inventory) > (page + 1)*10:
        nav.append(InlineKeyboardButton("➡️ Next", callback_data="next"))
    if nav:
        buttons.append(nav)

    buttons.append([
        InlineKeyboardButton("🛒 View Cart / Checkout", callback_data="checkout")
    ])

    if update.callback_query:
        await update.callback_query.edit_message_text("Select records to add to cart:",
            reply_markup=InlineKeyboardMarkup(buttons)
        )
    else:
        await update.message.reply_text("Select records to add to cart:",
            reply_markup=InlineKeyboardMarkup(buttons)
        )

    return SEARCH_VINYL

# ➕ Add to cart
async def handle_inventory_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if query.data == "next":
        context.user_data["page"] += 1
        return await show_inventory_page(update, context)
    elif query.data == "prev":
        context.user_data["page"] -= 1
        return await show_inventory_page(update, context)
    elif query.data == "checkout":
        return await show_cart(update, context)
    elif query.data.startswith("add_"):
        idx = int(query.data.split("_")[1])
        selected = context.user_data["inventory"][idx]
        context.user_data["cart"].append(selected)
        await query.answer("Added to cart ✅")
        return await show_inventory_page(update, context)

# 🛒 View cart
async def show_cart(update: Update, context: ContextTypes.DEFAULT_TYPE):
    cart = context.user_data["cart"]
    if not cart:
        await update.callback_query.answer("Cart is empty.")
        return SEARCH_VINYL

    text = "🛒 Cart:\n\n"
    for item in cart:
        text += f"• {item['Artist - Album']} - ${item['USD Price']}\n"

    buttons = [
        [InlineKeyboardButton("💳 Pay with POS", callback_data="pay_pos")],
        [InlineKeyboardButton("💵 Pay with Cash", callback_data="pay_cash")],
        [InlineKeyboardButton("🔙 Back", callback_data="back")]
    ]

    await update.callback_query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(buttons))
    return CONFIRM_CART

# 💰 Payment method
async def handle_payment(update: Update, context: ContextTypes.DEFAULT_TYPE):
    method = "POS" if "pos" in update.callback_query.data else "Cash"
    cart = context.user_data["cart"]
    for item in cart:
        save_sale(item, method)
        remove_record_from_inventory(item)
    await update.callback_query.edit_message_text(f"✅ {len(cart)} record(s) sold via {method}.\nInventory updated.")
    return ConversationHandler.END

# 📁 Save sale
def save_sale(record, payment_method):
    today = datetime.date.today().isoformat()
    path = os.path.join(SALES_FOLDER, f"{today}.xlsx")

    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Date", "Artist - Album", "Genre", "Style", "Label", "Format", "Condition", "USD Price", "Payment Method"])

    ws.append([
        today,
        record["Artist - Album"], record["Genre"], record["Style"],
        record["Label"], record.get("Format", "N/A"), record["Condition"],
        record["USD Price"], payment_method
    ])
    wb.save(path)

# 📉 Remove from inventory
def remove_record_from_inventory(target):
    from openpyxl import load_workbook, Workbook
    wb = load_workbook(INVENTORY_FILE)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    all_rows = list(ws.iter_rows(min_row=2, values_only=True))

    kept = []
    match = False
    for row in all_rows:
        row_dict = dict(zip(headers, row))
        if not match and all(str(row_dict.get(k)) == str(target.get(k)) for k in headers):
            match = True
            continue
        kept.append(row)

    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.append(headers)
    for row in kept:
        new_ws.append(row)
    new_wb.save(INVENTORY_FILE)

def start_sell_flow():
    return [ConversationHandler(
        entry_points=[CommandHandler("sell", start_sell)],
        states={
            SEARCH_VINYL: [CallbackQueryHandler(handle_inventory_selection)],
            CONFIRM_CART: [CallbackQueryHandler(handle_payment, pattern=r"^pay_"),
                           CallbackQueryHandler(show_inventory_page, pattern="^back$")]
        },
        fallbacks=[],
        name="sell_flow",
        persistent=False
    )]
